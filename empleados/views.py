from django.shortcuts import render, redirect,HttpResponse
from django.http import JsonResponse

#loggin
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.db import IntegrityError
from datetime import datetime

from .models import Producto

#historial de búsqueda
from .models import SearchHistory

from django.template.loader import render_to_string

from django.conf import settings

import tempfile

from weasyprint import HTML

#categoria
from .models import Categoria

import requests

import os
import uuid
from django.core.files.uploadedfile import SimpleUploadedFile

#bibliotecas para generar PDF
from reportlab.pdfgen import canvas 



from decimal import Decimal  # Asegúrate de importar Decimal
from django.contrib import messages  # Para usar mensajes flash
from django.core.exceptions import ObjectDoesNotExist

# carrito de compras
from empleados.Carrito import Carrito
from empleados.models import Producto, Empleado

# Para el informe (Reporte) Excel
import pandas as pd

import json

import logging
from django.contrib import messages
from django.utils import timezone
from openpyxl import Workbook  # Para generar el informe en excel
from django.http import HttpResponse, JsonResponse

from django.shortcuts import get_object_or_404
from . models import Empleado  # Importando el modelo de Empleado

from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone

from django.conf import settings
from django.core.paginator import Paginator
import io

import requests

import certifi

from xhtml2pdf import pisa

from django.db import transaction

from .models import Pedido, ItemPedido

from datetime import datetime
from calendar import monthrange

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

from openpyxl.styles import Font, PatternFill, Border, Side, numbers
from django.utils.timezone import make_aware, is_naive
from openpyxl.drawing.image import Image as ExcelImage
import zipfile
from tempfile import TemporaryDirectory, NamedTemporaryFile
from openpyxl.styles import Font, PatternFill, Border, Side, numbers
from django.core.files import File

from django.utils.timezone import make_aware, is_naive

from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.core.mail import send_mail
from django.shortcuts import render
from calendar import monthrange
from django.db.models import Sum
from collections import OrderedDict
from django.db.models.functions import ExtractMonth
from collections import defaultdict

from django.db.models import Sum
from collections import OrderedDict
import calendar
from .models import ItemPedido
from io import BytesIO
from django.views.decorators.http import require_POST
from django.db.models import Q


#anular pedido
@require_POST
def anular_pedido(request):
    try:
        data = json.loads(request.body)
        id_pedido = data.get('idPedido')
        pedido = get_object_or_404(Pedido, id=id_pedido)

        # Reponer stock
        for item in pedido.items.all():
            producto = item.producto
            producto.edad_empleado += item.cantidad  # ← aquí está el stock
            producto.save()


        # Eliminar el pedido (o marcar como anulado)
        pedido.delete()  # o pedido.estado = 'anulado'; pedido.save()

        return JsonResponse({'resultado': 1})
    except Exception as e:
        print("Error al anular pedido:", e)
        return JsonResponse({'resultado': 0})







#papelera de productos 
def papelera_productos(request):
    empleados = Empleado.objects.filter(activo=False)
    return render(request, 'empleado/papelera_empleados.html', {
        'empleados': empleados
    })


@require_POST
def restaurar_empleado(request):
    try:
        id_empleado = json.loads(request.body)['idEmpleado']
        empleado = get_object_or_404(Empleado, id=id_empleado)
        empleado.activo = True
        empleado.save()
        return JsonResponse({'resultado': 1})
    except Exception as e:
        print("Error al restaurar empleado:", e)
        return JsonResponse({'resultado': 0})



#ayuda

def como_comprar(request):
    return render(request, 'empleado/como_comprar.html')



#alerta de productos por acabarse 
def productos_alerta_view(request):
    sin_stock = Empleado.objects.filter(edad_empleado=0)
    bajo_stock = Empleado.objects.filter(edad_empleado__lte=5, edad_empleado__gt=0)

    context = {
        'sin_stock': sin_stock,
        'bajo_stock': bajo_stock,
    }
    return render(request, 'empleado/productos_alerta.html', context)




#informacion de la tienda
def info_tienda_view(request):
    return render(request, 'empleado/info_tienda.html')


#resumen de ventas 
def historial_general_compras(request):
    month = request.GET.get('month')  # Ej: "09"

    pedidos = Pedido.objects.all()  # 🔄 Aquí quitamos el filtro por usuario

    if month:
        try:
            month_int = int(month)
            year = datetime.now().year
            start_date = datetime(year, month_int, 1)
            end_day = monthrange(year, month_int)[1]
            end_date = datetime(year, month_int, end_day, 23, 59, 59)
            pedidos = pedidos.filter(fecha__range=(start_date, end_date))
        except ValueError:
            pass

    pedidos = pedidos.order_by('-fecha')

    meses = [
        ('01', 'Enero'), ('02', 'Febrero'), ('03', 'Marzo'), ('04', 'Abril'),
        ('05', 'Mayo'), ('06', 'Junio'), ('07', 'Julio'), ('08', 'Agosto'),
        ('09', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre')
    ]

    context = {
        'pedidos': pedidos,
        'selected_month': month,
        'meses': meses,
    }
    return render(request, 'empleado/historial_general.html', context)

Pedido


def exportar_excel(request):
    month = request.GET.get('month')  # Ej: "09"
    year = datetime.now().year 
    pedidos = Pedido.objects.all().select_related('usuario')

    if month:
        try:
            month_int = int(month)
            start_date = datetime(year, month_int, 1)
            end_day = monthrange(year, month_int)[1]
            end_date = datetime(year, month_int, end_day, 23, 59, 59)
            pedidos = pedidos.filter(fecha__range=(start_date, end_date))
        except ValueError:
            pass

    pedidos = pedidos.order_by('-fecha')

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen mensual"

    # Estilos
    bold_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Usuario", "Código", "Fecha", "USD", "Bs", "Tasa", "Método Pago", "Método Entrega"]
    ws.append(headers)

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font
        cell.fill = fill
        cell.border = border

    total_usd = 0
    total_bs = 0

    for pedido in pedidos:
        total_usd += float(pedido.monto_usd or 0)
        total_bs += float(pedido.monto_bs or 0)

        ws.append([
            pedido.usuario.username if pedido.usuario else "Sin usuario",
            pedido.codigo or "Sin código",
            pedido.fecha.strftime("%Y-%m-%d %H:%M") if pedido.fecha else "Sin fecha",
            float(pedido.monto_usd or 0),
            float(pedido.monto_bs or 0),
            float(pedido.tasa or 0),
            pedido.get_metodo_pago_display() or "Sin método",
            pedido.metodo_entrega or "Sin entrega"
        ])

        for item in pedido.items.all():
            ws.append([
                "",  # para no repetir pedido
                f"Producto: {item.producto.nombre_empleado if item.producto else 'Sin producto'}",
                f"Cantidad: {item.cantidad}",
                f"Subtotal: {item.subtotal}",
                "", "", "", "", ""
            ])

    # Fila vacía + totales
    ws.append([""] * 8)
    ws.append([
        "TOTAL", "", "", total_usd, total_bs, "", "", ""
    ])

    # Ajuste automático de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Guardar en memoria
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    mes_num = str(month).zfill(2) if month else datetime.now().strftime("%m")
    mes_nombre = {
        '01': 'enero', '02': 'febrero', '03': 'marzo', '04': 'abril',
        '05': 'mayo', '06': 'junio', '07': 'julio', '08': 'agosto',
        '09': 'septiembre', '10': 'octubre', '11': 'noviembre', '12': 'diciembre'
    }.get(mes_num, 'mes')




    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"resumen_{mes_nombre}_{year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'



    return response

def pedido_detalle(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    items = pedido.items.all()  # Esto ya funciona gracias al related_name
    return render(request, 'empleado/detalle.html', {'pedido': pedido, 'items': items})



def resumen_mensual_ventas(request):
    año_param = request.GET.get('year')
    try:
        año = int(año_param) if año_param else datetime.now().year
    except ValueError:
        año = datetime.now().year

    # 🔄 Agrupación manual por mes
    pedidos = Pedido.objects.filter(fecha__year=año, fecha__isnull=False)
    totales_por_mes = defaultdict(float)

    for p in pedidos:
        if p.fecha:
            mes = p.fecha.month
            totales_por_mes[mes] += float(p.monto_usd or 0)

    # 🧱 Diccionario base con todos los meses
    base_meses = {i: totales_por_mes.get(i, 0) for i in range(1, 13)}

    # 📊 Construimos resumen y datos para Chart.js
    resumen = []
    labels = []
    totales = []
    anterior = None

    for mes_num in range(1, 13):
        nombre_mes = calendar.month_name[mes_num]
        total = base_meses[mes_num]
        tendencia = None

        if anterior is not None:
            if total > anterior:
                tendencia = 'sube'
            elif total < anterior:
                tendencia = 'baja'
            else:
                tendencia = 'estable'

        resumen.append({
            'mes': nombre_mes,
            'total': total,
            'tendencia': tendencia
        })

        labels.append(nombre_mes)
        totales.append(total)
        anterior = total

    context = {
        'resumen': resumen,
        'labels': labels,
        'totales': totales,
        'año': año,
        'años_disponibles': list(range(2023, datetime.now().year + 1)),
    }

    return render(request, 'empleado/resumen_mensual.html', context)

def ranking_usuarios(request):
    usuarios = User.objects.annotate(total_compras=Sum('pedido__monto_usd'))\
                           .filter(total_compras__gt=0)\
                           .order_by('-total_compras')

    context = {'usuarios': usuarios}
    return render(request, 'empleado/ranking_usuarios.html', context)

def compras_por_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)
    pedidos = Pedido.objects.filter(usuario=usuario).order_by('-fecha').prefetch_related('items__producto')

    context = {
        'usuario': usuario,
        'pedidos': pedidos,
    }
    return render(request, 'empleado/compras_usuario.html', context)


def ranking_productos(request):
    ranking = (
        ItemPedido.objects
        .values('producto__nombre_empleado')
        .annotate(total_vendido=Sum('cantidad'))
        .order_by('-total_vendido')
    )
    return render(request, 'empleado/ranking_productos.html', {'ranking': ranking})








#logging

def signup(request):
    if request.method == 'GET':
        return render(request, 'empleado/signup.html')
    else:
        username = request.POST.get("username")
        email = request.POST.get("email")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if password1 != password2:
            return render(request, 'empleado/signup.html', {
                "error": "Las contraseñas no coinciden."
            })

        try:
            user = User.objects.create_user(username=username, email=email, password=password1)
            user.save()
            login(request, user)
            return redirect('listar_empleados')
        except IntegrityError:
            return render(request, 'empleado/signup.html', {
                "error": "El nombre de usuario ya está en uso."
            })
    
def signout(request):
    logout(request)
    return redirect('signin')

def signin(request):
    if request.method == 'GET':
        return render(request, 'empleado/signin.html', {"form": AuthenticationForm})
    else:
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'empleado/signin.html', {"form": AuthenticationForm, "error": "Username or password is incorrect."})

        login(request, user)
        return redirect('listar_empleados')



def password_reset_request(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        user = User.objects.filter(email=email).first()

        if user:
            # Aquí puedes generar un token o simplemente enviar el nombre de usuario
            send_mail(
                subject='Recuperación de cuenta',
                message=f'Tu nombre de usuario es: {user.username}',
                from_email='noreply@tuapp.com',
                recipient_list=[email],
                fail_silently=False,
            )
            return render(request, 'empleado/password_reset_sent.html', {'email': email})
        else:
            return render(request, 'empleado/password_reset_request.html', {
                'error': 'No se encontró ninguna cuenta con ese correo.'
            })

    return render(request, 'empleado/password_reset_request.html')

def recover_username(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        user = User.objects.filter(email=email).first()
        if user:
            send_mail(
                'Recuperación de usuario',
                f'Tu nombre de usuario es: {user.username}',
                DEFAULT_FROM_EMAIL,
                [email],
            )
            return render(request, 'empleado/username_sent.html', {'email': email})
        else:
            return render(request, 'empleado/recover_username.html', {'error': 'Correo no encontrado.'})
    return render(request, 'empleado/recover_username.html')




# Generación de PDF para la pago mobil

@transaction.atomic
def generar_pdf_pedido(request):
    if request.method == 'POST':
        data = request.POST
        carrito = request.session.get('carrito', {})
        monto_total = sum(float(item['acumulado']) for item in carrito.values())
        tasa = dataApiBcv['dollar']
        total_bolivares = round(monto_total * tasa, 2) if tasa else 0
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        codigo = request.POST.get('codigo', '')
        logo_path = os.path.join(settings.BASE_DIR, 'empleados', 'static', 'imgs', 'fotologo', 'fotosambilito.png')
        usuario = request.user.get_full_name() or request.user.username
        metodo_pago = request.POST.get('metodo_pago', 'no especificado')

        # 🔻 Descontar stock
        for item in carrito.values():
            try:
                producto = Empleado.objects.filter(nombre_empleado=item['nombre']).first()

                cantidad = int(item['cantidad'])

                if producto.edad_empleado >= cantidad:
                    producto.edad_empleado -= cantidad
                    producto.save()
                else:
                    return JsonResponse({'error': f'Stock insuficiente para {producto.nombre_empleado}'}, status=400)
            except Empleado.DoesNotExist:
                return JsonResponse({'error': f'Producto no encontrado: {item["nombre"]}'}, status=404)

        pedido = Pedido.objects.create(
            usuario=request.user,
            codigo=codigo,
            monto_usd=monto_total,
            monto_bs=total_bolivares,
            tasa=tasa,
            metodo_pago=metodo_pago,
            metodo_entrega=request.POST.get('metodo_entrega', 'no especificado'),
        )

        for item in carrito.values():
            producto = Empleado.objects.filter(nombre_empleado=item['nombre']).first()

            ItemPedido.objects.create(
                pedido=pedido,
                producto=producto,
                cantidad=int(item['cantidad']),
                subtotal=float(item['acumulado'])
            )


        context = {
            'pago_Movil': {'monto': monto_total},
            'total_ves': total_bolivares,
            'tasa': tasa,
            'carrito': carrito,
            'usuario': usuario,
            'fecha': fecha,
            'codigo': codigo,
            'logo_path': logo_path,
        }

        html = render_to_string('empleado/pedido_pdf.html', context)
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), dest=result)

        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="pedido_{codigo}.pdf"'
            # 🧹 Limpiar carrito si todo salió bien
            request.session['carrito'] = {}
            return response
        else:
            return JsonResponse({'error': 'Error al generar el PDF'}, status=500)


#usuario
def cuenta_view(request):
    user = request.user
    username_form = UsernameForm(instance=user)
    email_form = EmailForm(instance=user)
    password_form = PasswordChangeForm(user=user)

    if request.method == 'POST':
        if 'update_username' in request.POST:
            username_form = UsernameForm(request.POST, instance=user)
            if username_form.is_valid():
                username_form.save()
                messages.success(request, 'Nombre de usuario actualizado.')
                return redirect('cuenta')

        elif 'update_email' in request.POST:
            email_form = EmailForm(request.POST, instance=user)
            if email_form.is_valid():
                email_form.save()
                messages.success(request, 'Correo electrónico actualizado.')
                return redirect('cuenta')

        elif 'update_password' in request.POST:
            password_form = PasswordChangeForm(user=user, data=request.POST)
            if password_form.is_valid():
                password_form.save()
                update_session_auth_hash(request, password_form.user)
                messages.success(request, 'Contraseña actualizada.')
                return redirect('cuenta')

    return render(request, 'empleado/cuenta.html', {
        'user': user,
        'username_form': username_form,
        'email_form': email_form,
        'password_form': password_form,
    })

class UsernameForm(forms.ModelForm):
    class Meta:
        model = User
        fields = ['username']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control'}),
        }
class EmailForm(forms.ModelForm):
    class Meta:
        model = User
        fields = ['email']
        widgets = {
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
        }




#historial de compras
def historial_compras(request):
    month = request.GET.get('month')  # Ej: "09"

    pedidos = Pedido.objects.filter(usuario=request.user)

    if month:
        try:
            month_int = int(month)
            year = datetime.now().year  # Puedes ajustar esto si quieres filtrar por otro año
            start_date = datetime(year, month_int, 1)
            end_day = monthrange(year, month_int)[1]
            end_date = datetime(year, month_int, end_day, 23, 59, 59)
            pedidos = pedidos.filter(fecha__range=(start_date, end_date))
        except ValueError:
            pass

    pedidos = pedidos.order_by('-fecha')

    meses = [
        ('01', 'Enero'), ('02', 'Febrero'), ('03', 'Marzo'), ('04', 'Abril'),
        ('05', 'Mayo'), ('06', 'Junio'), ('07', 'Julio'), ('08', 'Agosto'),
        ('09', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre')
    ]

    context = {
        'pedidos': pedidos,
        'selected_month': month,
        'meses': meses,
    }
    return render(request, 'empleado/historial.html', context)






#genera pdf para efectivo
@transaction.atomic
def generar_pdf_efectivo(request):
    if request.method == 'POST':
        data = request.POST
        carrito = request.session.get('carrito', {})
        monto_total = sum(float(item['acumulado']) for item in carrito.values())
        tasa = dataApiBcv['dollar']
        total_bolivares = round(monto_total * tasa, 2) if tasa else 0
        fecha = timezone.now().strftime("%d/%m/%Y %H:%M")
        codigo = request.POST.get('codigo', '')
        metodo_pago = request.POST.get('metodo_pago', 'no especificado')
        logo_path = os.path.join(settings.BASE_DIR, 'empleados', 'static', 'imgs', 'fotologo', 'fotosambilito.png')
        usuario = request.user.get_full_name() or request.user.username

        # 🧾 Guardar el pedido en la base de datos
        pedido = Pedido.objects.create(
            usuario=request.user,
            codigo=codigo,
            monto_usd=monto_total,
            monto_bs=total_bolivares,
            tasa=tasa,
           metodo_pago=metodo_pago,
           
        )

        for item in carrito.values():
            producto = Empleado.objects.filter(nombre_empleado=item['nombre']).first()

            ItemPedido.objects.create(
                pedido=pedido,
                producto=producto,
                cantidad=int(item['cantidad']),
                subtotal=float(item['acumulado'])
            )
            # Descontar stock
            producto.edad_empleado -= int(item['cantidad'])
            producto.save()

        context = {
            'metodo_pago': metodo_pago,  # ← esto sí refleja lo que se envió
            'monto': monto_total,
            'total_ves': total_bolivares,
            'tasa': tasa,
            'carrito': carrito,
            'usuario': usuario,
            'fecha': fecha,
            'codigo': codigo,
            'logo_path': logo_path,
        }

        html = render_to_string('empleado/efectivo_pdf.html', context)
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), dest=result)

        if not pdf.err:
            filename = f'pedido_{codigo}.pdf'
            pdf_path = os.path.join(settings.MEDIA_ROOT, 'facturas', filename)
            with open(pdf_path, 'wb') as f:
                f.write(result.getvalue())

            # 🧹 Vaciar el carrito
            request.session['carrito'] = {}

            return JsonResponse({'pdf_url': f'/media/facturas/{filename}'})




#metodo pago


def obtener_dolar_bcv():
    url = 'https://bcv-api.rafnixg.dev/rates'  # Puedes cambiarla por otra API que permita CORS
    try:
        response = requests.get(url)
        response.raise_for_status()  # Lanza excepción si hay error HTTP
        data = response.json()
        return data  # O guarda todo el JSON si lo necesitas
    except requests.RequestException as e:
        print(f"Error al obtener el dólar BCV: {e}")
        return "No disponible"  # Manejo de errores, puedes personalizarlo
dataApiBcv = obtener_dolar_bcv()


#vsita de pago
def pago_movil(request):
    carrito = request.session.get('carrito', {})
    monto_total = sum(float(item['acumulado']) for item in carrito.values())
    tasa = dataApiBcv['dollar']
    if tasa is None:
        messages.error(request, "No se pudo obtener la tasa del dólar. Intente más tarde.")
        total_bolivares = 0
    else:
        total_bolivares = round(monto_total * tasa, 2)

    pagoMovil={
        'documento': '31419483',
        'telefono': '04247726393',
        'monto': monto_total,         # Total en dólares
        'concepto': 'codigo123',
    }

    contexto = {
        'pago_Movil': pagoMovil,
        'total_ves': total_bolivares,
        'tasa': tasa,
        'carrito_json': json.dumps(carrito)
    }
    return render(request, 'empleado/pago_movil.html', contexto)


def efectivo(request):
    carrito = request.session.get('carrito', {})
    monto_total = sum(float(item['acumulado']) for item in carrito.values())
    tasa = dataApiBcv['dollar']
    total_bolivares = round(monto_total * tasa, 2) if tasa else 0

    contexto = {
        'pago_Movil': {
            'monto': monto_total
        },
        'total_ves': total_bolivares,
        'tasa': tasa,
        'carrito_json': json.dumps(carrito)
    }
    return render(request, 'empleado/efectivo.html', contexto)



#buscador de productos 


def buscar_productos(request):
    query = request.GET.get('query', '')
    resultados = Empleado.objects.filter(
        Q(nombre_empleado__icontains=query) |
        Q(apellido_empleado__icontains=query) |
        Q(categoria__nombre__icontains=query),
        activo=True
    )

    # Guardar búsqueda si el usuario está autenticado
    if request.user.is_authenticated and query:
        SearchHistory.objects.create(user=request.user, query=query)

    return render(request, 'empleado/resultado_busqueda.html', {
        'query': query,
        'resultados': resultados
    })



#historial de búsqueda

def search_history_view(request):
    query = request.GET.get('q', '')
    month = request.GET.get('month', '')

    history_list = SearchHistory.objects.filter(user=request.user)

    if query:
        history_list = history_list.filter(query__icontains=query)

    if month:
        try:
            month_int = int(month)
            year = datetime.now().year
            start_date = datetime(year, month_int, 1)
            end_day = monthrange(year, month_int)[1]
            end_date = datetime(year, month_int, end_day, 23, 59, 59)
            history_list = history_list.filter(timestamp__range=(start_date, end_date))
        except ValueError:
            pass

    history_list = history_list.order_by('-timestamp')
    paginator = Paginator(history_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    meses = [
        ('01', 'Enero'), ('02', 'Febrero'), ('03', 'Marzo'), ('04', 'Abril'),
        ('05', 'Mayo'), ('06', 'Junio'), ('07', 'Julio'), ('08', 'Agosto'),
        ('09', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre'),
    ]

    return render(request, 'empleado/historial_busqueda.html', {
        'page_obj': page_obj,
        'query': query,
        'selected_month': month,
        'meses': meses,
    })



def eliminar_busqueda(request, pk):
    busqueda = get_object_or_404(SearchHistory, pk=pk, user=request.user)
    busqueda.delete()
    return redirect('search_history')

def eliminar_todo_historial(request):
    if request.method == 'POST':
        SearchHistory.objects.filter(user=request.user).delete()
        messages.success(request, 'Historial de búsqueda eliminado correctamente.')
    return redirect('search_history') 


#carrito de compras
from empleados.models import Empleado

def carrito_view(request):
    carrito = Carrito(request)
    carrito_data = request.session.get('carrito', {})
    total_carrito = sum(item['acumulado'] for item in carrito_data.values())
    tasa = dataApiBcv['dollar']
    total_bs = round(total_carrito * tasa, 2)
    empleados = Empleado.objects.filter(id__in=[item['empleado_id'] for item in carrito_data.values()])

    contexto = {
        'total_bs': total_bs,
        'tasa': tasa,
        'carrito': carrito,
        'total_carrito': total_carrito,
        'empleados': {str(e.id): e for e in empleados} # ← diccionario para acceso rápido
    }

    return render(request, 'empleado/carrito.html', contexto)


def agregar_producto(request, empleado_id):
    if not request.user.is_authenticated:
        return JsonResponse({
            'error': 'Debes iniciar sesión para agregar productos al carrito.'
        }, status=401)

    carrito = Carrito(request)
    empleado = get_object_or_404(Empleado, id=empleado_id)
    try:
        carrito.agregar(empleado)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)

    carrito_data = request.session.get('carrito', {})
    empleados = Empleado.objects.filter(id__in=[item['empleado_id'] for item in carrito_data.values()])

    html = render_to_string('empleado/_carrito_contenido.html', {
        'carrito_data': carrito_data,
        'total_carrito': sum(i.get('acumulado', 0) for i in carrito_data.values()),
        'empleados': {str(e.id): e for e in empleados}
    }, request)

    return JsonResponse({'html': html})





def sumar_producto_ajax(request, key):
    key = str(key)
    carrito = request.session.get('carrito', {})

    if key not in carrito:
        return JsonResponse({'error': 'Producto no encontrado en carrito'}, status=404)

    item = carrito[key]
    empleado = get_object_or_404(Empleado, id=int(key))

    # Validación de stock
    if item['cantidad'] >= empleado.edad_empleado:
        return JsonResponse({'error': 'No puedes agregar más unidades que las disponibles en stock.'}, status=400)

    precio = float(item.get('precio', empleado.salario_empleado) or 0)
    cantidad = item.get('cantidad', 0) + 1

    item['cantidad'] = cantidad
    item['precio'] = precio
    item['acumulado'] = round(cantidad * precio, 2)
    carrito[key] = item

    request.session['carrito'] = carrito
    request.session.modified = True

    carrito_data = carrito
    empleados = Empleado.objects.filter(id__in=[item['empleado_id'] for item in carrito_data.values()])
    html = render_to_string('empleado/_carrito_contenido.html', {
        'carrito_data': carrito_data,
        'total_carrito': sum(i.get('acumulado', 0) for i in carrito_data.values()),
        'empleados': {str(e.id): e for e in empleados}
    }, request)

    return JsonResponse({'html': html})


def restar_producto_ajax(request, key):
    key = str(key)
    carrito = request.session.get('carrito', {})

    if key not in carrito:
        return JsonResponse({'error': 'Producto no encontrado en carrito'}, status=404)

    item = carrito[key]
    precio = float(item.get('precio', 0) or 0)
    cantidad = item.get('cantidad', 0) - 1

    if cantidad <= 0:
        del carrito[key]
    else:
        item['cantidad'] = cantidad
        item['acumulado'] = round(cantidad * precio, 2)
        carrito[key] = item

    request.session['carrito'] = carrito
    request.session.modified = True

    carrito_data = carrito
    empleados = Empleado.objects.filter(id__in=[item['empleado_id'] for item in carrito_data.values()])
    html = render_to_string('empleado/_carrito_contenido.html', {
        'carrito_data': carrito_data,
        'total_carrito': sum(i.get('acumulado', 0) for i in carrito_data.values()),
        'empleados': {str(e.id): e for e in empleados}
    }, request)

    return JsonResponse({'html': html})


def eliminar_producto_ajax(request, key):
    key = str(key)
    carrito = request.session.get('carrito', {})

    if key in carrito:
        del carrito[key]
        request.session['carrito'] = carrito
        request.session.modified = True

    carrito_data = carrito
    empleados = Empleado.objects.filter(id__in=[item['empleado_id'] for item in carrito_data.values()])
    html = render_to_string('empleado/_carrito_contenido.html', {
        'carrito_data': carrito_data,
        'total_carrito': sum(i.get('acumulado', 0) for i in carrito_data.values()),
        'empleados': {str(e.id): e for e in empleados}
    }, request)

    return JsonResponse({'html': html})




def limpiar_carrito(request):
    carrito = Carrito(request)
    carrito.limpiar()
    return redirect('listar_empleados')





#vista efectivo
def vista_efectivo(request):
    return render(request, 'empleado/efectivo.html')


#productos

def inicio(request):
    categoria_id = request.GET.get('categoria')
    categorias = Categoria.objects.all()

    if categoria_id:
        empleados = Empleado.objects.filter(categoria_id=categoria_id, activo=True)
    else:
        empleados = Empleado.objects.filter(activo=True)

    tasa = dataApiBcv['dollar'] 
    return render(request, 'empleado/lista_empleados.html', {
        'empleados': empleados,
        'categorias': categorias,
        'categoria_seleccionada': categoria_id,
        'tasa': tasa
    })





def listar_categorias(request):
    categorias = Categoria.objects.all()
    return render(request, 'empleado/form_empleado.html', {'categorias': categorias})



def listar_empleados(request):
    categoria_id = request.GET.get('categoria')
    categorias = Categoria.objects.all()

    if categoria_id:
        empleados = Empleado.objects.filter(categoria_id=categoria_id, activo=True)
    else:
        empleados = Empleado.objects.filter(activo=True)

    tasa = dataApiBcv['dollar'] 
    return render(request, 'empleado/lista_empleados.html', {
        'empleados': empleados,
        'categorias': categorias,
        'categoria_seleccionada': categoria_id,
        'tasa': tasa
    })

def view_form_carga_masiva(request):
    return render(request, 'empleado/form_carga_masiva.html')


def detalles_empleado(request, id):
    try:
        empleado = Empleado.objects.get(id=id)
        data = {"empleado": empleado}
        return render(request, "empleado/detalles.html", data)
    except Empleado.DoesNotExist:
        error_message = f"no existe ningún registro para la busqueda id: {id}"
        return render(request, "empleado/lista_empleados.html", {"error_message": error_message})


def registrar_empleado(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre_empleado', '').strip()
        apellido = request.POST.get('apellido_empleado', '').strip()
        edad = request.POST.get('edad_empleado', '').strip()
        salario_raw = request.POST.get('salario_empleado', '').strip()
        categoria_id = request.POST.get('categoria')
        foto_empleado = request.FILES.get('foto_empleado')

        # 🔍 Validación del salario
        try:
            salario = Decimal(salario_raw.replace(',', '.'))
        except (InvalidOperation, AttributeError):
            messages.error(request, "El salario ingresado no es válido. Usa punto como separador decimal.")
            return redirect('view_form_carga_masiva')  # o la vista que corresponda

        # 🔍 Verificar si ya existe un empleado con ese apellido
        producto_existente = Empleado.objects.filter(apellido_empleado=apellido).first()

        if producto_existente:
            foto_path = None
            if foto_empleado:
                foto_path = default_storage.save(f"empleados/{foto_empleado.name}", foto_empleado)
            request.session['nuevo_producto'] = {
                'nombre': nombre,
                'apellido': apellido,
                'edad': edad,
                'salario': str(salario),
                'categoria_id': categoria_id,
                'foto_path': foto_path
            }
            return redirect('confirmar_reemplazo', producto_id=producto_existente.id)

        # 🧱 Crear el nuevo empleado
        try:
            categoria = Categoria.objects.get(id=categoria_id)
        except Categoria.DoesNotExist:
            messages.error(request, "La categoría seleccionada no existe.")
            return redirect('view_form_carga_masiva')

        empleado = Empleado(
            nombre_empleado=nombre,
            apellido_empleado=apellido,
            edad_empleado=edad,
            salario_empleado=salario,
            categoria=categoria
        )
        if foto_empleado:
            empleado.foto_empleado = foto_empleado

        empleado.save()
        messages.success(request, f"Felicitaciones, el producto {nombre} fue registrado correctamente 😉")
        return redirect('listar_empleados')

    categorias = Categoria.objects.all()
    return render(request, 'empleado/form_empleado.html', {'categorias': categorias})

def confirmar_reemplazo(request, producto_id):
    producto_existente = Empleado.objects.filter(id=producto_id).first()
    nuevo = request.session.get('nuevo_producto')
    if not nuevo:
        messages.error(request, "No se encontraron datos para reemplazar.")
        return redirect('registrar_empleado')

    return render(request, 'empleado/confirmar_reemplazo.html', {
        'producto': producto_existente,
        'nuevo': nuevo
    })

def reemplazar_producto(request, producto_id):
    if request.method == 'POST':
        producto = Empleado.objects.get(id=producto_id)
        nuevo = request.session.get('nuevo_producto')

        if nuevo:
            producto.nombre_empleado = nuevo['nombre']
            producto.apellido_empleado = nuevo.get('apellido', producto.apellido_empleado)
            producto.edad_empleado = nuevo.get('edad', producto.edad_empleado)
            producto.salario_empleado = nuevo.get('salario', producto.salario_empleado)

            # Reemplazar categoría
            try:
                categoria = Categoria.objects.get(id=nuevo['categoria_id'])
                producto.categoria = categoria
            except Categoria.DoesNotExist:
                pass  # Puedes mostrar un mensaje si lo prefieres

            # Reemplazar imagen si hay una nueva
            if nuevo.get('foto_path'):
                producto.foto_empleado = nuevo['foto_path']

            producto.save()
            del request.session['nuevo_producto']
            messages.success(request, "Empleado reemplazado correctamente.")
        else:
            messages.error(request, "No se pudo completar el reemplazo.")

        return redirect('listar_empleados')


#actualizar producto
def view_form_update_empleado(request, id):
    try:
        empleado = Empleado.objects.get(id=id)
        opciones_edad = [(edad, edad) for edad in range(18, 51)]
        categorias = Categoria.objects.all()

        return render(request, "empleado/form_update_empleado.html", {
            "empleado": empleado,
            "opciones_edad": opciones_edad,
            "categorias": categorias,
        })
    except Empleado.DoesNotExist:
        error_message = f"El Empleado con id: {id} no existe."
        return render(request, "empleado/lista_empleados.html", {"error_message": error_message})



def actualizar_empleado(request, id):
    empleado = get_object_or_404(Empleado, id=id)

    if request.method == 'POST':
        empleado.nombre_empleado = request.POST.get('nombre_empleado')
        empleado.apellido_empleado = request.POST.get('apellido_empleado')
        empleado.edad_empleado = request.POST.get('edad_empleado')
        empleado.salario_empleado = request.POST.get('salario_empleado')
        raw_salario = request.POST.get('salario_empleado', '0')
        salario = raw_salario.replace(',', '.')
        empleado.salario_empleado = salario


        categoria_id = request.POST.get('categoria')
        if categoria_id:
            try:
                empleado.categoria = Categoria.objects.get(id=categoria_id)
            except Categoria.DoesNotExist:
                pass

        if 'foto_empleado' in request.FILES:
            empleado.foto_empleado = request.FILES['foto_empleado']

        empleado.save()
        return redirect('listar_empleados')  # o la vista que prefieras

    return redirect('view_form_update_empleado', id=id)

#descarga de productos en excel
def informe_empleado(request):
    try:
        with TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'data_empleados.xlsx')
            imagenes_dir = os.path.join(temp_dir, 'imagenes')
            os.makedirs(imagenes_dir, exist_ok=True)

            # Crear Excel
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Informe de Empleados"
            worksheet.freeze_panes = "A2"

            headers = [
                'Nombre del Producto',
                'Especificación del Producto',
                'Categoría del Producto',
                'Cantidad del Producto',
                'Precio del Producto',
                'Fecha sin zona horaria',
                'Foto'
            ]
            worksheet.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            empleados = Empleado.objects.select_related('categoria').all()

            for empleado in empleados:
                fecha = empleado.created_at.replace(tzinfo=None) if empleado.created_at else ''
                nombre_imagen = os.path.basename(empleado.foto_empleado.name) if empleado.foto_empleado else ''

                worksheet.append([
                    empleado.nombre_empleado,
                    empleado.apellido_empleado,
                    empleado.categoria.nombre if empleado.categoria else '',
                    empleado.edad_empleado,
                    empleado.salario_empleado,
                    fecha,
                    nombre_imagen
                ])

                row = worksheet.max_row
                worksheet.cell(row=row, column=5).number_format = '"$"#,##0.00'
                worksheet.cell(row=row, column=6).number_format = numbers.FORMAT_DATE_DATETIME

                # Copiar imagen al directorio temporal
                if empleado.foto_empleado and os.path.exists(empleado.foto_empleado.path):
                    destino = os.path.join(imagenes_dir, nombre_imagen)
                    imagen = PILImage.open(empleado.foto_empleado.path)

                    # Convertir si tiene canal alfa (transparencia)
                    if imagen.mode == 'RGBA':
                        imagen = imagen.convert('RGB')

                    imagen.save(destino)


            # Ajuste de ancho de columnas
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

            workbook.save(excel_path)

            # Crear ZIP
            zip_path = os.path.join(temp_dir, 'empleados_con_imagenes.zip')
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                zipf.write(excel_path, arcname='data_empleados.xlsx')
                for img_name in os.listdir(imagenes_dir):
                    img_path = os.path.join(imagenes_dir, img_name)
                    zipf.write(img_path, arcname=f'imagenes/{img_name}')

            # Descargar ZIP
            with open(zip_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="empleados_con_imagenes.zip"'
                return response

    except Exception as e:
        return HttpResponse(f"Error al generar el ZIP: {str(e)}", content_type="text/plain") 


#eliminar producto

def eliminar_empleado(request):
    try:
        id_empleado = json.loads(request.body)['idEmpleado']
        empleado = get_object_or_404(Empleado, id=id_empleado)
        empleado.activo = False  # ← Aquí desactivamos el producto
        empleado.save()
        return JsonResponse({'resultado': 1})
    except Exception as e:
        print("Error al desactivar empleado:", e)
        return JsonResponse({'resultado': 0})


#carga masiva de productos

logger = logging.getLogger(__name__)

def cargar_archivo(request):
    if request.method != 'POST':
        return JsonResponse({'status_server': 'error', 'message': 'Método HTTP no válido.'})

    archivo_zip = request.FILES.get('archivo_zip')
    if not archivo_zip or not archivo_zip.name.endswith('.zip'):
        return JsonResponse({'status_server': 'error', 'message': 'Debes subir un archivo ZIP válido.'})

    try:
        import pandas as pd

        errores = []
        nuevos = 0
        actualizados = 0

        with TemporaryDirectory() as temp_dir:
            zip_path = os.path.join(temp_dir, archivo_zip.name)
            with open(zip_path, 'wb') as f:
                for chunk in archivo_zip.chunks():
                    f.write(chunk)

            with zipfile.ZipFile(zip_path, 'r') as zipf:
                zipf.extractall(temp_dir)

            # Buscar archivo Excel
            excel_file = None
            for file in os.listdir(temp_dir):
                if file.endswith('.xlsx'):
                    excel_file = os.path.join(temp_dir, file)
                    break

            if not excel_file:
                return JsonResponse({'status_server': 'error', 'message': 'No se encontró ningún archivo Excel (.xlsx) dentro del ZIP.'})

            df = pd.read_excel(excel_file, header=0)

            columnas_requeridas = [
                'Nombre del Producto',
                'Especificación del Producto',
                'Categoría del Producto',
                'Cantidad del Producto',
                'Precio del Producto',
                'Fecha sin zona horaria',
                'Foto'
            ]
            columnas_encontradas = df.columns.tolist()
            faltantes = [col for col in columnas_requeridas if col not in columnas_encontradas]

            if faltantes:
                logger.error(f"Columnas faltantes: {faltantes}")
                return JsonResponse({'status_server': 'error', 'message': f'Faltan columnas requeridas: {", ".join(faltantes)}'})

            for index, row in df.iterrows():
                try:
                    nombre = str(row['Nombre del Producto']).strip()
                    apellido = str(row['Especificación del Producto']).strip()
                    categoria_nombre = str(row['Categoría del Producto']).strip()
                    cantidad = int(row['Cantidad del Producto'])
                    precio = float(row['Precio del Producto'])
                    fecha = row['Fecha sin zona horaria']

                    if isinstance(fecha, datetime) and is_naive(fecha):
                        fecha = make_aware(fecha)

                    categoria_obj, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)

                    empleado, creado = Empleado.objects.update_or_create(
                        nombre_empleado=nombre,
                        apellido_empleado=apellido,
                        categoria=categoria_obj,
                        defaults={
                            'edad_empleado': cantidad,
                            'salario_empleado': precio,
                            'created_at': fecha,
                        }
                    )

                    # Buscar imagen en carpeta extraída
                    foto_nombre = str(row['Foto']).strip()
                    foto_path = os.path.join(temp_dir, 'imagenes', foto_nombre)
                    if not os.path.exists(foto_path):
                        foto_path = os.path.join(temp_dir, foto_nombre)

                    if foto_nombre and os.path.exists(foto_path):
                        with open(foto_path, 'rb') as f:
                            empleado.foto_empleado.save(foto_nombre, File(f), save=True)

                    if creado:
                        nuevos += 1
                    else:
                        actualizados += 1

                except Exception as fila_error:
                    logger.error(f"Error en fila {index + 2}: {fila_error}")
                    errores.append(f"Fila {index + 2}: {fila_error}")

        mensaje_final = f"Importación completada. Nuevos: {nuevos}, Actualizados: {actualizados}."
        status = 'success' if not errores else 'partial'

        return JsonResponse({
            'status_server': status,
            'message': mensaje_final,
            'errores': errores
        })

    except Exception as e:
        logger.error(f"Error al procesar el ZIP: {e}")
        return JsonResponse({
            'status_server': 'error',
            'message': f'Error al procesar el archivo ZIP: {str(e)}'   })
        




# Genera un nombre único para el archivo utilizando UUID y conserva la extensión.
def generate_unique_filename(file):
    extension = os.path.splitext(file.name)[1]
    unique_name = f'{uuid.uuid4()}{extension}'
    return SimpleUploadedFile(unique_name, file.read())
